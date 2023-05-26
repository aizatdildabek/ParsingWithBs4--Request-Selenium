from bs4 import BeautifulSoup
import requests
import openpyxl
import re
from time import sleep

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'SmartphoneAlser07.3.2023'
print(excel.sheetnames)
sheet.append(['name_tovar', 'price', 'countSimcard', 'dioganal', 'screen_resolution', 'techno_screen', 'opSystem', 'count_yader', 'ssd', 'datchik'])

for cnt in range(1,22):
    my_list=[]
    url=f'https://alser.kz/c/vse-smartfony?page={cnt}'
    responce=requests.get(url)

    soup=BeautifulSoup(responce.text, 'html.parser')
    #print(soup)


    data=soup.find_all('div', class_="col-md-3")
    #print(data)    
    for i in data:
        if i.find('a', class_='product-item__info_title')!=None:
            name=i.find('a', class_='product-item__info_title').text
        if i.find('div', class_='price')!=None:
            price=i.find('div', class_='price').text
        
        if i.find_all('div', class_='row product-card-spec')!=None:
            description = i.find_all('div', class_='row product-card-spec')
            if len(description)>0:
                my_list.clear()
                for x in range(len(description)):
                    desc=description[x].text
                    count="".join(re.findall(r'- ([^<>]+)', desc))
                    my_list.append(count)
                print(name, price, *my_list, end=' ')
                sheet.append([name, price, *my_list])




    excel.save('SmartphoneAlser07.03.2023.xlsx')        





