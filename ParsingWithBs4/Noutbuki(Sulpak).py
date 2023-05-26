from bs4 import BeautifulSoup
import requests
import openpyxl
from time import sleep
import re


excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'Noutbuki'
print(excel.sheetnames)
sheet.append(['Название товара', 'Диагональ','Процессор','Операционная система' ,'Объем оперативной памяти','Тип жесткого диска','Объем накопителя','Серия видеокарты', 'Чипсет видеоадаптера','Объем памяти видеоадаптера', 'Цена', 'Наличие товара'])


responce=requests.get(url='https://www.sulpak.kz/f/noutbuki', headers={'User-Agent': 'python-requests/2.28.2', 'Accept-Encoding': 'gzip, deflate', 'Accept': '*/*', 'Connection': 'keep-alive', 'one': 'true'})

for count in range(1, 18):
    sleep(3)
    url=f"https://www.sulpak.kz/f/noutbuki?page={count}" 

    responce=requests.get(url)

    soup=BeautifulSoup(responce.text,"html.parser")

    # print(soup)

    data=soup.find_all("div", class_="product__item product__item-js tile-container")

    # print(data)

    for i in data:

        name=i.find("div", class_="product__item-name").text

        description=i.find("div", class_="product__item-description").text

        diagonal="".join(re.findall(r'Диагональ:([^<>]+)/ Процессор', description))

        processor="".join(re.findall(r'Процессор:([^<>]+)/ Операционная система', description))

        opSystem = "".join(re.findall(r'Операционная система:([^<>]+)/ Объем оперативной памяти', description))

        ram= "".join(re.findall(r'Объем оперативной памяти:([^<>]+)/ Тип жесткого диска', description)) 

        disk="".join(re.findall(r'Тип жесткого диска:([^<>]+)/ Объем накопителя', description)) 

        Storage_capacity="".join(re.findall(r'Объем накопителя:([^<>]+)/ Серия видеокарты', description)) 

        videocard="".join(re.findall(r'Серия видеокарты:([^<>]+)/ Чипсет видеоадаптера', description)) 

        CHvideoadapter="".join(re.findall(r'Чипсет видеоадаптера:([^<>]+)/ Объем памяти видеоадаптера', description)) 

        Ovideoadapter="".join(re.findall(r'Объем памяти видеоадаптера:([^<>]+)', description)) 

        if i.find("div", class_="product__item-price")==None:
            Price=0
        else:
            Price= i.find("div", class_="product__item-price").text   
  
        if i.find("div", class_="product__item-showcase")==None:
            product_availability='Нет нового товара, есть уцененный'
        else:    
            product_availability=i.find("div", class_="product__item-showcase").text


        print(name, diagonal,  processor, opSystem, ram, disk, Storage_capacity, videocard, CHvideoadapter, Ovideoadapter, Price, product_availability)
        sheet.append([name, diagonal,  processor, opSystem, ram, disk, Storage_capacity, videocard, CHvideoadapter, Ovideoadapter, Price, product_availability])


    excel.save('Noutbuki.xlsx')