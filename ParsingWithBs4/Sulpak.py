from bs4 import BeautifulSoup
import requests
import openpyxl
from time import sleep
import re


excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'Smartphone'
print(excel.sheetnames)
sheet.append(['Название товара', 'Операционная система','Количество SIM-карт','Диагональ дисплея' ,'Объем встроенной памяти','Основная камера','Фронтальная камера','NFC','Старые цена', 'Скидка', 'Новые Цена', 'Наличие товара'])

for count in range(1,31):
    sleep(3)
    url=f"https://www.sulpak.kz/f/smartfoniy?page={count}" 

    responce=requests.get(url)

    soup=BeautifulSoup(responce.text,"html.parser")

    # print(soup)

    data=soup.find_all("div", class_="product__item product__item-js tile-container")

    #print(data)

    for i in data:

        name=i.find("div", class_="product__item-name").text

        description=i.find("div", class_="product__item-description").text

        opSystem = "".join(re.findall(r'Операционная система:([^<>]+)/ Количество', description))

        countSIMcard="".join(re.findall(r'Количество SIM-карт:([^<>]+)/ Диагональ', description)) 

        display="".join(re.findall(r'Диагональ дисплея:([^<>]+)/ Объем', description)) 

        ram= "".join(re.findall(r'Объем встроенной памяти:([^<>]+)/ Основная', description)) 

        camera="".join(re.findall(r'Основная камера:([^<>]+)/ Фронтальная', description)) 

        frontCamera="".join(re.findall(r'Фронтальная камера:([^<>]+)/ NFC', description)) 

        nfc="".join(re.findall(r'NFC:([^<>]+)', description)) 

        if i.find("div", class_="product__item-price-old")==None:

            oldPrice=""
        else:
            oldPrice1=i.find("div", class_="product__item-price-old").text  
            oldPrice="".join(re.findall(r'([^<>]+)₸', oldPrice1))   

        if i.find("div", class_="product__label-discount")==None:
           
            sale=""
        else:
            sale=i.find("div", class_="product__label-discount").text

        if i.find("div", class_="product__item-price")==None:
            newPrice=""
        else:
            newPrice1=i.find("div", class_="product__item-price").text
            newPrice= "".join(re.findall(r'([^<>]+)₸', newPrice1))  
  
        if i.find("div", class_="product__item-showcase")==None:
            if i.find("div", class_="product__item-stock-empty")==None:
                product_availability="Нет нового товара, есть уцененный"
            else:
                product_availability=i.find("div", class_="product__item-stock-empty").text    
        else:    
            product_availability=i.find("div", class_="product__item-showcase").text

        print(name, opSystem, countSIMcard, display, ram, camera, frontCamera, nfc, oldPrice, sale, newPrice, product_availability)
        sheet.append([name, opSystem, countSIMcard, display, ram, camera, frontCamera, nfc, oldPrice, sale, newPrice, product_availability])


    excel.save('Smartphones.xlsx')