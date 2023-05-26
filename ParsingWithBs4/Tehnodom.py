from bs4 import BeautifulSoup
import requests
import openpyxl
from time import sleep
import re

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'SmartphoneTechnodom07.03.2023'
print(excel.sheetnames)
sheet.append(['name_tovar', 'price', 'year', 'dioganal', 'display', 'type_matr', 'chastata_obnovlenie', 'operative_disk', 'bstroennoy_disk', 'model_processor','opSystem', 'interface', 'standart_security', 'weight'])

for count in range(1,5):
    my_list=[]
    url=f'https://www.technodom.kz/catalog/smartfony-i-gadzhety/smartfony-i-telefony?page={count}'

    resp=requests.get(url)

    soup=BeautifulSoup(resp.text,"html.parser")

    data=soup.find_all('li', class_='category-page-list__item')


    for i in data:
        if i.find('p', class_='Typography ProductCardV_card__title__MK4_q ProductCardV_--loading__2C9Aq Typography__M')==None:
            name=''
        else:
            name=i.find('p', class_='Typography ProductCardV_card__title__MK4_q ProductCardV_--loading__2C9Aq Typography__M').text
        # print(name)
        if i.find('p', class_='Typography ProductCardPrices_prices-info__price__y8O_n Typography__Subtitle')==None:
            price=''
        else:    
            price=i.find('p', class_='Typography ProductCardPrices_prices-info__price__y8O_n Typography__Subtitle').text
        # print(price)
        if i.find('a', class_='category-page-list__item-link')!=None:   
            href='https://www.technodom.kz'+i.find("a", class_='category-page-list__item-link')['href']

            resp1=requests.get(href)
            soup1=BeautifulSoup(resp1.text,"html.parser")

            if soup1.find_all('p',class_='Typography product-description__right-text Typography__Body Typography__Body_Small')!=None:
                description = soup1.find_all('p', class_='Typography product-description__right-text Typography__Body Typography__Body_Small')
                if len(description)>0:
                    my_list.clear()
                    for x in range(len(description)):
                        desc=description[x].text
                        my_list.append(desc)
                    print(name, price, *my_list, end=' ')
                    sheet.append([name, price, *my_list])

    excel.save('SmartphoneTechnodom07.03.2023.xlsx')               