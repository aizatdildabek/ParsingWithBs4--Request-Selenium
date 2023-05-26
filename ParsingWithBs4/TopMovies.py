from bs4 import BeautifulSoup
import requests
import openpyxl


excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'Top Rated Movies'
print(excel.sheetnames)
sheet.append(['Movie Rank', 'Movie Name', 'Year of Release', 'Rating'])

#get запросить какой то информацию
#post отправить информацию

# response=requests.get("https://httpbin.org")
# print(response.status_code)     #статус код 200 возвращает когда запрос успешно, а 404 означает что такой адрес не найдена
# print(response.headers)
# print(response.content)   #b в байтовай виде, значит двоичном строке//сре данные-необработанные данны
# print(response.text)     #об данные
# print(response.json())  #ответ в Jsone


source = requests.get('https://www.imdb.com/chart/top/')
source.raise_for_status()
    
soup = BeautifulSoup(source.text, 'html.parser')
print(soup) # возвращает html код
# movies = soup.find('tbody', class_="lister-list").find_all('tr')
# #print(len(movies)) #250 row, because 250 top film
# for movie in movies:

#     name = movie.find('td',class_='titleColumn').a.text  #html tag <a>
#    #rank = movie.find('td', class_='titleColumn').get_text(strip=True)
#     rank = movie.find('td', class_='titleColumn').get_text(strip=True).split('.')[0]  #[0]-index 0 возвращает только номер фильма
#     year = movie.find('td', class_='titleColumn').span.text.strip('()') #возвращает только год премьера/strip возвращает без ()
#     rating = movie.find('td', class_='ratingColumn imdbRating').strong.text  #strong html tag 
#     print(rank, name, year, rating)
#     sheet.append([rank, name, year, rating])

# excel.save('MovieRatings.xlsx')