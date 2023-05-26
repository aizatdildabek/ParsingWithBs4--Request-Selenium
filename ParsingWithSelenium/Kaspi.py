from selenium.webdriver import Chrome
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait as wait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import openpyxl
from time import sleep
import re
# from tqdm import tqdm #var заполняемость прогресса


excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'Kaspi'
print(excel.sheetnames)
sheet.append(['Name', 'Price', 'NFC', 'color','type_screen','dioganal' ,'RAM_size','Processor','memory_capacity','Battery'])


browser=Chrome('/Users/user/Desktop/chromedriver')
for count in range(100, 120):
    sleep(3)
    url=f'https://kaspi.kz/shop/search/?text=%D0%B2%D1%81%D0%B5%20%D1%81%D0%BC%D0%B0%D1%80%D1%82%D1%84%D0%BE%D0%BD%D1%8B&page={count}'

    browser.get(url)

    # browser.find_element(By.CLASS_NAME, 'search-bar__input')
    # input_tub=browser.find_element(By.CLASS_NAME, 'search-bar__input')
    # input_tub.send_keys('все смартфоны')
    # btn=browser.find_element(By.CLASS_NAME, 'search-bar__submit')
    # btn.click()
    my_list=[]

    soup=BeautifulSoup(browser.page_source, 'html.parser')

    data=soup.find_all('div', class_='item-card ddl_product ddl_product_link undefined')
    
    for i in data:
        name=i.find('a', class_='item-card__name-link').text.strip()
        price=i.find('span', class_='item-card__prices-price').text
        # print(name)
        # print(price)
        href=i.find('a', class_='item-card__image-wrapper')['href']
        # print(href)
        browser.get(href)
        soup1=BeautifulSoup(browser.page_source, 'html.parser')
        description=soup1.find_all('li', class_='short-specifications__text')
        my_list.clear()

        for x in range(len(description)):
            desc=description[x].text
            count="".join(re.findall(r': ([^<>]+)', desc))
            my_list.append(count) 
        print(name, price, *my_list, end=' ')    
        sheet.append([name, price, *my_list])


    excel.save('KaspiDatasets1.xlsx')    